#!/usr/bin/env python3
"""
fetch_rbnz_nzd.py — RBNZ B2 "10 year" govt bond closing yield → data/NZD_NOM.csv

Source: hb2-daily-close.xlsx (the live workbook since RBNZ moved B2 to closing
rates with a 1-business-day lag in 2025; the old hb2-daily.xlsx is discontinued).

Output CSV shape (consumed by the dashboard's parseNomCsv):
    # QUALITY=CLEAN source=RBNZ_B2_hb2-daily-close asof=<UTC ts>
    DATE,NOM10
    2026-06-13,4.46
    ...

Design notes (institutional robustness):
  * The 10Y column is located by HEADER TEXT, not a hard-coded index — RBNZ
    re-benchmarks bonds periodically (e.g. May-2036 issue) and column order can
    shift. We match a header containing both a "10" and a year/bond cue.
  * Excel date serials AND real datetimes are both handled.
  * If the workbook can't be fetched or the 10Y column can't be found, the
    script EXITS NON-ZERO and writes NOTHING — the dashboard then stays on its
    manual fallback rather than ingesting a malformed file.
  * Keeps the trailing ~600 business days to bound file size.
"""

import io
import sys
import datetime as dt

import requests
from openpyxl import load_workbook

URL = ("https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/"
       "statistics/series/b/b2/hb2-daily-close.xlsx")
OUT = "data/NZD_NOM.csv"
UA = ("Mozilla/5.0 (g8-macro-pipeline; "
      "+https://github.com/sanderdayan1982/g8-macro-pipeline)")
KEEP = 600   # trailing rows to retain


def die(msg, code=1):
    print(f"[fetch_rbnz_nzd] FATAL: {msg}", file=sys.stderr)
    sys.exit(code)


def fetch_workbook():
    try:
        r = requests.get(URL, headers={"User-Agent": UA}, timeout=60)
    except Exception as e:
        die(f"request failed: {e}")
    if r.status_code != 200:
        die(f"HTTP {r.status_code} from RBNZ (ASN may be blocked on this runner)")
    if len(r.content) < 5000:
        die(f"suspiciously small payload ({len(r.content)} bytes) — likely an error page")
    return io.BytesIO(r.content)


def find_10y_column(ws):
    """Scan the first ~12 rows for a header cell that denotes the 10-year
    benchmark govt bond yield. Returns (header_row_idx, col_idx) 1-based."""
    cues_strong = ["10 year", "10-year", "10yr", "10 yr"]
    for ri in range(1, 13):
        for ci in range(1, ws.max_column + 1):
            val = ws.cell(row=ri, column=ci).value
            if val is None:
                continue
            s = str(val).strip().lower()
            if not s:
                continue
            # strong textual match
            if any(c in s for c in cues_strong) and ("govt" in s or "government"
                    in s or "bond" in s or "benchmark" in s or s.startswith("10")):
                return ri, ci
            # RBNZ series-id style, e.g. a "10" benchmark column header
            if s in ("10", "10y") :
                return ri, ci
    return None, None


def coerce_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)):
        # Excel serial (1900 date system)
        try:
            return (dt.date(1899, 12, 30) + dt.timedelta(days=int(v)))
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y", "%Y%m%d"):
            try:
                return dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def coerce_value(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace("%", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def main():
    buf = fetch_workbook()
    try:
        wb = load_workbook(buf, read_only=True, data_only=True)
    except Exception as e:
        die(f"openpyxl could not read workbook: {e}")

    # B2 close workbook: the data sheet is usually "Data" — fall back to scan.
    sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() in ("data", "daily", "b2 daily", "hb2"):
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[-1]]   # last sheet is typically the data sheet

    hdr_row, col = find_10y_column(sheet)
    if col is None:
        die("could not locate the 10-year column by header — RBNZ may have "
            "changed the workbook layout; inspect hb2-daily-close.xlsx")

    # Find the date column: first column whose header looks like a date label,
    # else default to column 1.
    date_col = 1
    for ci in range(1, sheet.max_column + 1):
        hv = sheet.cell(row=hdr_row, column=ci).value
        if hv and str(hv).strip().lower() in ("date", "series id", "series_id"):
            date_col = ci
            break

    rows = []
    for ri in range(hdr_row + 1, sheet.max_row + 1):
        d = coerce_date(sheet.cell(row=ri, column=date_col).value)
        y = coerce_value(sheet.cell(row=ri, column=col).value)
        if d is None or y is None:
            continue
        # sanity band: NZ 10Y has lived roughly within [-1, 12]% historically
        if -1.0 <= y <= 12.0:
            rows.append((d, y))

    if not rows:
        die("no parseable (date,value) rows found in the 10-year column")

    rows.sort(key=lambda t: t[0])
    rows = rows[-KEEP:]

    asof = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%MZ")
    lines = [f"# QUALITY=CLEAN source=RBNZ_B2_hb2-daily-close asof={asof}",
             "DATE,NOM10"]
    for d, y in rows:
        lines.append(f"{d.isoformat()},{y:.3f}")

    with open(OUT, "w", newline="\n") as f:
        f.write("\n".join(lines) + "\n")

    print(f"[fetch_rbnz_nzd] wrote {OUT}: {len(rows)} rows, "
          f"last {rows[-1][0].isoformat()} = {rows[-1][1]:.3f}%")


if __name__ == "__main__":
    main()

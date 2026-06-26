"""
fetch_gbp_bills.py
==================
Scraper for UK gilt nominal spot yields from Bank of England.

Source:   Bank of England — UK Government Liability Curve, nominal (two ZIPs, merged)
Endpoints:
    1. LATEST (fresh, daily):
       https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip
       Small ZIP (~400 KB). THIS is the source of fresh data. Published by noon
       on the following business day.

       PACKAGING NOTE (changed ~2026-06): the BoE repackaged this endpoint. The
       outer ZIP no longer holds the workbooks directly — it now holds preview
       GIFs (uknom/ukinf/ukois/ukreal.gif) plus a NESTED ZIP
       'Latest Yield Curve data (current month).zip' that contains the four
       current-month workbooks (Nominal/Inflation/Real/OIS). We descend one
       level of nesting to reach the nominal workbook. The flat layout is still
       attempted first, so a future revert keeps working.

    2. HISTORICAL (stable base):
       https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/glcnominalddata.zip
       Large ZIP (~38 MB). Contains GLC_Nominal_daily_data_<years>.xlsx files
       split by year range. NOTE: this archive is refreshed only periodically
       (observed to lag by weeks), so it must NOT be the sole source. It
       provides the stable multi-year history; the latest ZIP provides
       freshness.

Why both:
    Using the 38 MB archive alone freezes the data at the archive's last
    server-side refresh. The latest ZIP carries the live current-month tail.
    Merging the two yields full 5-year history AND fresh data, deduplicating on
    date with the latest file winning on any overlap.

Freshness gate (added 2026-06):
    The merge can SILENTLY fall back to the stale archive if the LATEST source
    breaks (as happened when BoE nested the ZIP: the parser failed, was caught,
    and the run still reported "5 OK"). A "5 OK" with month-old data must not
    pass as success. main() now checks the age of the newest 10Y point and
    returns a non-zero exit (failing the CI step) if it exceeds
    STALENESS_LIMIT_DAYS — so a broken upstream surfaces the same day, not weeks
    later.

Anti-cache:
    BoE FAQ warns that a cached page can serve stale spreadsheets. We send
    Cache-Control/Pragma no-cache headers plus a cache-busting query param to
    force fresh copies.

Sheets in each XLSX:
    info                — documentation/disclaimer
    1. fwds, short end  — forward rates short end (months)
    2. fwd curve        — forward rates full curve
    3. spot, short end  — spot rates short end (months 1-60)
    4. spot curve       — SPOT RATES FULL CURVE (0.5Y, 1Y, 1.5Y, ...)  <- we use this

Structure of sheet "4. spot curve" (verified identical in both latest & archive):
    Row 0: title ("UK nominal spot curve")
    Row 1: blank
    Row 2: "Maturity"
    Row 3: "years:", 0.5, 1, 1.5, 2, 2.5, 3, ...   <- maturity in years
    Row 4: "#VALUE!" placeholder — skip
    Row 5+: datetime in col 0, yields in subsequent columns
    (trailing all-None rows at end — skipped defensively)

Why 6M as bill_short_GBP:
    The BoE gilt nominal curve has 0.5-year (=6M) as its shortest stable point.
    This matches DGS6MO in fetch_us_bills.py, giving GBP a symmetric 6M/6M pair
    vs US (no curve-slope bias). No 3M point exists in this curve.

Output:
    data/GBP_BILL_6M.csv   — 6-month gilt nominal (bill_short_GBP for engine)
    data/GBP_BILL_1Y.csv   — 1-year
    data/GBP_BILL_2Y.csv   — 2-year
    data/GBP_BILL_5Y.csv   — 5-year   (curve cross-validation)
    data/GBP_BILL_10Y.csv  — 10-year  (benchmark)

License: Bank of England public statistics. Per BoE FAQ "Can I store the data
         provided on your website?": electronic storage permitted per their
         legal/disclaimer page. Citation: Bank of England, UK Government
         Liability Curve — Nominal (daily).
"""

import csv
import io
import re
import sys
import time
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import requests


# Constants
BOE_LATEST_URL = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/"
    "yield-curves/latest-yield-curve-data.zip"
)
BOE_ARCHIVE_URL = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/"
    "yield-curves/glcnominalddata.zip"
)
HISTORY_YEARS = 5
TIMEOUT_SECONDS = 120  # archive ZIP is ~38 MB
USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

# Max age (calendar days) of the newest 10Y point before the run is treated as
# a failure. ~7 days ≈ 4-5 business days, matching the dashboard DQM budget and
# covering a long bank-holiday weekend without false alarms.
STALENESS_LIMIT_DAYS = 7

# Which sheet inside each XLSX holds the spot curve we want
SPOT_CURVE_SHEET = "4. spot curve"

# Tenor in years -> output filename. 0.5 = 6 months.
TENORS = {
    0.5: "GBP_BILL_6M.csv",   # bill_short_GBP for engine
    1.0: "GBP_BILL_1Y.csv",
    2.0: "GBP_BILL_2Y.csv",
    5.0: "GBP_BILL_5Y.csv",
    10.0: "GBP_BILL_10Y.csv",
}

# In the ARCHIVE zip, files are named GLC_Nominal_daily_data_<start>_to_<end>.xlsx
ARCHIVE_XLSX_PATTERN = re.compile(
    r"GLC[ _]Nominal[ _]daily[ _]data[ _](\d{4})[ _]to[ _](\d{4}|present)\.xlsx$",
    re.IGNORECASE,
)
# In the LATEST zip (flat layout), the nominal file is named exactly:
LATEST_NOMINAL_NAME = "GLC Nominal daily data current month.xlsx"

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data"


def _select_relevant_archive_names(
    zip_namelist: List[str],
    earliest_year_needed: int,
) -> List[str]:
    """From the archive ZIP, return only XLSX whose year range overlaps the window."""
    selected: List[str] = []
    current_year = datetime.now(timezone.utc).year
    for name in zip_namelist:
        m = ARCHIVE_XLSX_PATTERN.search(name)
        if not m:
            continue
        start = int(m.group(1))
        end_raw = m.group(2)
        end = current_year if end_raw == "present" else int(end_raw)
        if end >= earliest_year_needed and start <= current_year:
            selected.append(name)
    return selected


def _find_nominal_xlsx(names: List[str]) -> Optional[str]:
    """
    Pick the nominal current-month workbook from a ZIP namelist.

    Exact name first (old flat layout), then a defensive fuzzy match: any .xlsx
    whose name contains 'nominal'. This deliberately ignores the .gif previews
    (uknom.gif etc.) and the Inflation/Real/OIS siblings, which do not contain
    the substring 'nominal'.
    """
    if LATEST_NOMINAL_NAME in names:
        return LATEST_NOMINAL_NAME
    candidates = [
        n for n in names
        if n.lower().endswith(".xlsx") and "nominal" in n.lower()
    ]
    return candidates[0] if candidates else None


def _extract_latest_nominal_bytes(latest_zip_bytes: bytes) -> bytes:
    """
    Return the bytes of the nominal current-month XLSX from the LATEST zip,
    descending one level of nesting if BoE has wrapped the workbooks in an inner
    ZIP (observed 2026-06). Raises ValueError with the discovered contents if it
    cannot be found, so the next repackaging is a one-shot debug.
    """
    with zipfile.ZipFile(io.BytesIO(latest_zip_bytes)) as zf:
        names = zf.namelist()

        # (a) flat layout: the workbook sits directly in the outer ZIP
        target = _find_nominal_xlsx(names)
        if target:
            print(f"    Latest layout: flat -> {target}")
            return zf.read(target)

        # (b) nested layout: open any inner .zip and look inside it
        nested = [n for n in names if n.lower().endswith(".zip")]
        seen_inner: List[str] = []
        for inner_name in nested:
            try:
                with zipfile.ZipFile(io.BytesIO(zf.read(inner_name))) as izf:
                    inner_names = izf.namelist()
                    seen_inner.extend(inner_names)
                    inner_target = _find_nominal_xlsx(inner_names)
                    if inner_target:
                        print(f"    Latest layout: nested -> {inner_name} -> {inner_target}")
                        return izf.read(inner_target)
            except zipfile.BadZipFile:
                continue

        raise ValueError(
            "Nominal current-month XLSX not found in latest ZIP. "
            f"Outer: {names}. Nested contents: {seen_inner or 'none'}"
        )


def _find_year_column_indices(
    header_row: tuple,
    wanted_tenors: List[float],
    tolerance: float = 1e-6,
) -> Dict[float, int]:
    """Map each wanted tenor (in years) to its column index from the 'years:' header row."""
    col_for_tenor: Dict[float, int] = {}
    for j, cell in enumerate(header_row):
        if cell is None or isinstance(cell, str):
            continue
        try:
            cell_year = float(cell)
        except (TypeError, ValueError):
            continue
        for tenor in wanted_tenors:
            if tenor in col_for_tenor:
                continue
            if abs(cell_year - tenor) < tolerance:
                col_for_tenor[tenor] = j
    return col_for_tenor


def _parse_xlsx_bytes(
    xlsx_bytes: bytes,
    date_from: datetime,
    date_to: datetime,
) -> Dict[float, Dict[str, float]]:
    """
    Parse one BoE XLSX (in-memory) and extract rows for wanted tenors, filtered
    to [date_from, date_to]. Returns {tenor: {date_str: value}}.
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), read_only=True, data_only=True
    )
    if SPOT_CURVE_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Expected sheet '{SPOT_CURVE_SHEET}' not found. Got: {wb.sheetnames}"
        )

    ws = wb[SPOT_CURVE_SHEET]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 6:
        return {t: {} for t in TENORS}

    # Locate the 'years:' header row (normally row index 3)
    header_row = None
    for r in all_rows[:8]:
        if r and r[0] == "years:":
            header_row = r
            break
    if header_row is None:
        raise ValueError(f"Could not locate 'years:' header row in '{SPOT_CURVE_SHEET}'")

    col_for_tenor = _find_year_column_indices(header_row, list(TENORS.keys()))
    missing = [t for t in TENORS if t not in col_for_tenor]
    if missing:
        raise ValueError(
            f"Tenors {missing} not found in header. Sample: {header_row[:15]}"
        )

    out: Dict[float, Dict[str, float]] = {t: {} for t in TENORS}

    for raw in all_rows:
        if not raw or raw[0] is None:
            continue
        cell0 = raw[0]
        if isinstance(cell0, datetime):
            date_obj = cell0
        else:
            # Skip non-date rows (title/header/#VALUE! placeholder)
            try:
                date_obj = datetime.strptime(str(cell0).strip()[:10], "%Y-%m-%d")
            except ValueError:
                continue
        if date_obj < date_from or date_obj > date_to:
            continue
        date_str = date_obj.strftime("%Y%m%d")
        for tenor, col_idx in col_for_tenor.items():
            if col_idx >= len(raw):
                continue
            cell = raw[col_idx]
            if cell is None or cell == "" or isinstance(cell, str):
                continue
            try:
                value = float(cell)
            except (TypeError, ValueError):
                continue
            out[tenor][date_str] = value

    return out


def _fetch_zip(url: str, label: str, cache_bust: bool) -> bytes:
    """Download a ZIP with anti-cache headers. Returns raw bytes."""
    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "application/zip,*/*",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
    }
    params = {}
    if cache_bust:
        params["_"] = str(int(time.time()))
    print(f"  Fetching {label}: {url}")
    r = requests.get(url, headers=headers, params=params, timeout=TIMEOUT_SECONDS)
    r.raise_for_status()
    print(f"    Downloaded {len(r.content) / 1024:.0f} KB")
    return r.content


def fetch_gbp_bills(
    date_from: datetime,
    date_to: datetime,
) -> Dict[float, List[Tuple[str, float]]]:
    """
    Fetch GBP gilt yields from BOTH BoE endpoints and merge.

    Strategy:
        1. Parse archive ZIP (stable multi-year base).
        2. Parse latest ZIP current-month file (fresh tail), descending into a
           nested ZIP if the workbook is wrapped.
        3. Merge per tenor per date; latest wins on overlap.
        4. Return sorted lists.

    Resilience:
        If one source fails, return the other (degraded but not broken). Only
        raise if BOTH fail. NOTE: a successful merge from the archive alone is
        still STALE — main()'s freshness gate is what catches that.
    """
    archive_data: Dict[float, Dict[str, float]] = {t: {} for t in TENORS}
    latest_data: Dict[float, Dict[str, float]] = {t: {} for t in TENORS}
    archive_ok = False
    latest_ok = False

    # 1. Archive (stable base) — no cache-bust, it's large/static
    try:
        archive_bytes = _fetch_zip(BOE_ARCHIVE_URL, "archive (glcnominalddata, ~38MB)", cache_bust=False)
        with zipfile.ZipFile(io.BytesIO(archive_bytes)) as zf:
            relevant = _select_relevant_archive_names(zf.namelist(), date_from.year)
            if not relevant:
                raise ValueError(f"No relevant XLSX in archive. Names: {zf.namelist()}")
            print(f"    Reading {len(relevant)} archive file(s):")
            for name in sorted(relevant):
                print(f"      - {name.split('/')[-1]}")
                partial = _parse_xlsx_bytes(zf.open(name).read(), date_from, date_to)
                for tenor, dvals in partial.items():
                    archive_data[tenor].update(dvals)
        archive_ok = True
        ac = sum(len(v) for v in archive_data.values())
        print(f"    Archive parsed: {ac} (tenor,date) points")
    except (requests.RequestException, ValueError, zipfile.BadZipFile) as e:
        print(f"  WARNING: archive fetch/parse failed: {e}", file=sys.stderr)

    # 2. Latest (fresh tail) — cache-bust ON, descends into nested ZIP if needed
    try:
        latest_bytes = _fetch_zip(BOE_LATEST_URL, "latest (current month, ~400KB)", cache_bust=True)
        nominal_xlsx = _extract_latest_nominal_bytes(latest_bytes)
        partial = _parse_xlsx_bytes(nominal_xlsx, date_from, date_to)
        for tenor, dvals in partial.items():
            latest_data[tenor].update(dvals)
        latest_ok = True
        lc = sum(len(v) for v in latest_data.values())
        print(f"    Latest parsed: {lc} (tenor,date) points")
    except (requests.RequestException, ValueError, zipfile.BadZipFile) as e:
        print(f"  WARNING: latest fetch/parse failed: {e}", file=sys.stderr)

    if not archive_ok and not latest_ok:
        raise ValueError("Both BoE endpoints failed — no GBP data available")

    # 3. Merge: archive base, latest overlays (latest wins)
    merged: Dict[float, Dict[str, float]] = {t: {} for t in TENORS}
    for tenor in TENORS:
        merged[tenor].update(archive_data.get(tenor, {}))
        merged[tenor].update(latest_data.get(tenor, {}))

    # 4. Sorted lists
    results: Dict[float, List[Tuple[str, float]]] = {}
    for tenor in TENORS:
        results[tenor] = sorted(merged[tenor].items(), key=lambda kv: kv[0])

    return results


def write_csv(rows: List[Tuple[str, float]], output_path: Path) -> None:
    """Write rows to OHLCV format CSV (O=H=L=C for daily rates, V=0)."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["DATE", "OPEN", "HIGH", "LOW", "CLOSE", "VOLUME"])
        for date_str, value in rows:
            v = f"{value:.4f}"
            writer.writerow([date_str, v, v, v, v, "0"])


def main() -> int:
    today = datetime.now(timezone.utc).replace(tzinfo=None)
    date_from = today - timedelta(days=365 * HISTORY_YEARS)

    print(f"Fetching GBP gilt yields from {date_from.date()} to {today.date()}")
    print(f"Source: Bank of England (latest current-month + historical archive, merged)")
    print(f"Tenors: {[f'{int(t)}Y' if t >= 1 else f'{int(t*12)}M' for t in TENORS]}")
    print()

    try:
        results = fetch_gbp_bills(date_from, today)
    except requests.HTTPError as exc:
        print(f"ERROR: BoE HTTP error: {exc}", file=sys.stderr)
        return 1
    except requests.RequestException as exc:
        print(f"ERROR: BoE network error: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"ERROR: GBP bills fetch failed: {exc}", file=sys.stderr)
        return 1

    print()
    successes = 0
    failures = 0

    for tenor, filename in TENORS.items():
        rows = results.get(tenor, [])
        output_path = OUTPUT_DIR / filename
        label = f"{int(tenor*12)}M" if tenor < 1 else f"{int(tenor)}Y"

        if not rows:
            print(f"[{label}] ERROR: No rows returned", file=sys.stderr)
            failures += 1
            continue

        write_csv(rows, output_path)
        print(f"[{label}] OK: Wrote {len(rows)} rows to {filename}")
        print(f"        Latest:   {rows[-1][0]} = {rows[-1][1]:.4f}%")
        print(f"        Earliest: {rows[0][0]} = {rows[0][1]:.4f}%")
        successes += 1

    print()
    print(f"Summary: {successes} OK, {failures} failed (of {len(TENORS)} total)")

    # ── Freshness gate ──────────────────────────────────────────────────────
    # A merge that succeeded from the stale archive alone (because the LATEST
    # current-month source broke) is NOT a success. Fail loudly so CI goes red
    # the same day instead of weeks later. The 10Y is the benchmark series.
    stale = False
    benchmark = results.get(10.0, [])
    if benchmark:
        newest = benchmark[-1][0]
        try:
            age_days = (today - datetime.strptime(newest, "%Y%m%d")).days
        except ValueError:
            age_days = None
        if age_days is None:
            print(f"FRESHNESS WARN: could not parse newest 10Y date '{newest}'.", file=sys.stderr)
        elif age_days > STALENESS_LIMIT_DAYS:
            stale = True
            print(
                f"FRESHNESS FAIL: newest 10Y gilt is {newest} ({age_days}d old, "
                f"limit {STALENESS_LIMIT_DAYS}d). The LATEST current-month source did "
                f"not contribute — data merged from the historical archive only "
                f"(silent-staleness condition). Failing the run so it surfaces now. "
                f"Likely cause: BoE changed the latest-yield-curve-data.zip packaging.",
                file=sys.stderr,
            )
        else:
            print(f"Freshness OK: newest 10Y gilt {newest} ({age_days}d old, limit {STALENESS_LIMIT_DAYS}d).")

    return 0 if (failures == 0 and not stale) else 1


if __name__ == "__main__":
    sys.exit(main())

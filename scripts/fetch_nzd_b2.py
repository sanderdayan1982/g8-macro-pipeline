"""
fetch_nzd_b2.py — RBNZ Table B2 (Wholesale interest rates) daily close

Sources NZD short-end and government bond yields from a single RBNZ XLSX:
  - BKBM 30/60/90D bank bill yields (NZD analogue of BBSW/SOFR-bills)
  - NZ Government Bonds 1Y/2Y/5Y/10Y constant maturity closing yields

Endpoint:
  https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/statistics/series/b/b2/hb2-daily-close.xlsx

Outputs (one CSV per series, Date,Value with YYYY-MM-DD dates):
  NZD_BILL_30D.csv, NZD_BILL_60D.csv, NZD_BILL_90D.csv
  NZD_BOND_1Y.csv, NZD_BOND_2Y.csv, NZD_BOND_5Y.csv, NZD_BOND_10Y.csv

Frequency: daily, T-1 lag, released by RBNZ around 17:00 NZT.

Column matching is done by Series Id (row 5 of the Data sheet), NOT by column
position — this is resilient to RBNZ reordering columns in future releases.
"""

import os
import sys
import time
from io import BytesIO

# curl_cffi impersonates the TLS fingerprint of a real Chrome browser, which
# is required to bypass RBNZ's WAF. The default `requests` library exposes a
# Python/OpenSSL TLS signature (JA3 hash) that modern WAFs flag as bot traffic,
# triggering a 403 Forbidden even with browser-like User-Agent headers.
from curl_cffi import requests
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

URL = (
    "https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/"
    "statistics/series/b/b2/hb2-daily-close.xlsx"
)

# RBNZ's WAF applies TLS fingerprinting (likely Cloudflare) — it blocks any
# client whose TLS handshake doesn't match a real browser, regardless of
# User-Agent header. curl_cffi's impersonate="chrome" replicates Chrome's
# JA3 fingerprint + HTTP/2 settings + headers in a single call.
IMPERSONATE = "chrome"

OUTPUT_DIR = "data"
TIMEOUT_SECONDS = 90
RETRY_DELAYS = [0, 5, 15]  # seconds between attempts (3 attempts total)

# Series Id → output filename. Authoritative mapping.
SERIES_MAP = {
    # BKBM bank bill yields
    "INM.DB01.NZZV":   "NZD_BILL_30D.csv",
    "INM.DB02.NZZV":   "NZD_BILL_60D.csv",
    "INM.DB03.NZZV":   "NZD_BILL_90D.csv",
    # NZ Gov Bond constant maturity closing yields
    "INM.DG101.NZZCF": "NZD_BOND_1Y.csv",
    "INM.DG102.NZZCF": "NZD_BOND_2Y.csv",
    "INM.DG105.NZZCF": "NZD_BOND_5Y.csv",
    "INM.DG110.NZZCF": "NZD_BOND_10Y.csv",
}

SHEET_NAME = "Data"
SERIES_ID_ROW = 5  # row index (1-based) holding the Series Id labels
DATA_START_ROW = 6  # first row of observations
DATE_COL = 1        # column A holds dates

# ---------------------------------------------------------------------------
# Download with retry
# ---------------------------------------------------------------------------

def download_xlsx(url: str) -> bytes:
    """Fetch the XLSX bytes with exponential-style retry on transient errors.

    Uses curl_cffi with Chrome TLS impersonation to bypass RBNZ's WAF, which
    blocks non-browser TLS fingerprints (Python/OpenSSL signature) on requests
    coming from datacenter IPs like GitHub Actions runners.
    """
    last_exc = None
    for attempt, delay in enumerate(RETRY_DELAYS, start=1):
        if delay:
            print(f"[fetch_nzd_b2] retry in {delay}s...")
            time.sleep(delay)
        try:
            print(f"[fetch_nzd_b2] attempt {attempt}/{len(RETRY_DELAYS)} GET {url}")
            r = requests.get(url, timeout=TIMEOUT_SECONDS, impersonate=IMPERSONATE)
            r.raise_for_status()
            print(f"[fetch_nzd_b2] downloaded {len(r.content):,} bytes")
            return r.content
        except requests.exceptions.Timeout as e:
            last_exc = e
            print(f"[fetch_nzd_b2] timeout: {e}")
            continue
        except requests.exceptions.ConnectionError as e:
            last_exc = e
            print(f"[fetch_nzd_b2] connection error: {e}")
            continue
        except requests.exceptions.HTTPError as e:
            # Do not retry on HTTP errors (403, 404, 5xx server-side issues
            # that won't be cured by waiting a few seconds). If we get 403
            # here it means impersonation isn't enough — escalate to ops.
            print(f"[fetch_nzd_b2] HTTP error, aborting: {e}", file=sys.stderr)
            raise
    raise RuntimeError(
        f"[fetch_nzd_b2] download failed after {len(RETRY_DELAYS)} attempts: {last_exc}"
    )

# ---------------------------------------------------------------------------
# Parse XLSX
# ---------------------------------------------------------------------------

def locate_series_columns(ws) -> dict[str, int]:
    """Scan the Series Id row and return {series_id: column_index (1-based)} for targets.

    Uses iter_rows() so it works under read_only mode where ws.max_column is unreliable.
    """
    found: dict[str, int] = {}
    for row in ws.iter_rows(
        min_row=SERIES_ID_ROW, max_row=SERIES_ID_ROW, values_only=False
    ):
        for cell in row:
            if isinstance(cell.value, str) and cell.value in SERIES_MAP:
                found[cell.value] = cell.column  # 1-based column index

    missing = set(SERIES_MAP.keys()) - set(found.keys())
    if missing:
        raise RuntimeError(
            f"[fetch_nzd_b2] missing Series Ids in XLSX (RBNZ schema changed?): "
            f"{sorted(missing)}"
        )
    return found

def extract_series(ws, col_map: dict[str, int]) -> dict[str, list[tuple[str, float]]]:
    """Walk data rows and pull (date, value) pairs for each target series.

    Rows where the value is None/blank are skipped (NZ holidays, missing prints).
    """
    out: dict[str, list[tuple[str, float]]] = {sid: [] for sid in col_map}
    # Pre-compute (sid, zero_based_index) pairs to avoid dict lookups inside hot loop
    targets = [(sid, col - 1) for sid, col in col_map.items()]
    date_idx = DATE_COL - 1

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row:
            continue
        date_val = row[date_idx]
        if date_val is None:
            continue
        # RBNZ stores dates as datetime; coerce to YYYY-MM-DD
        try:
            date_str = date_val.strftime("%Y-%m-%d")
        except AttributeError:
            continue  # skip rows with unparseable date

        for sid, idx in targets:
            if idx >= len(row):
                continue
            v = row[idx]
            if v is None or v == "":
                continue
            try:
                out[sid].append((date_str, float(v)))
            except (TypeError, ValueError):
                continue  # skip non-numeric stray cells
    return out

# ---------------------------------------------------------------------------
# Write CSVs
# ---------------------------------------------------------------------------

def write_csv(path: str, rows: list[tuple[str, float]]) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write("Date,Value\n")
        for date_str, value in rows:
            f.write(f"{date_str},{value}\n")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    try:
        content = download_xlsx(URL)
    except Exception as e:
        print(f"[fetch_nzd_b2] FATAL download: {e}", file=sys.stderr)
        return 1

    try:
        wb = load_workbook(BytesIO(content), data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(
                f"sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}"
            )
        ws = wb[SHEET_NAME]

        col_map = locate_series_columns(ws)
        print(f"[fetch_nzd_b2] located all {len(col_map)} target series")

        data = extract_series(ws, col_map)
    except Exception as e:
        print(f"[fetch_nzd_b2] FATAL parse: {e}", file=sys.stderr)
        return 2

    # Write outputs
    total_rows = 0
    for sid, rows in data.items():
        filename = SERIES_MAP[sid]
        path = os.path.join(OUTPUT_DIR, filename)
        write_csv(path, rows)
        latest_date, latest_val = rows[-1] if rows else ("NONE", float("nan"))
        print(
            f"[fetch_nzd_b2] {filename}: {len(rows):,} rows, "
            f"latest {latest_date} = {latest_val}"
        )
        total_rows += len(rows)

    print(f"[fetch_nzd_b2] OK — {len(data)} files, {total_rows:,} total rows")
    return 0

if __name__ == "__main__":
    sys.exit(main())

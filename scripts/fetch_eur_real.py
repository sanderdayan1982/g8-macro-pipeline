#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fetch_eur_real.py
=================
EUR (núcleo alemán) REAL 10Y diario, real de MERCADO, desde los linkers del Bundesbank.

QUÉ HACE
--------
La Bundesbank publica cada día hábil la tabla "Prices and yields of listed Federal
securities": un XLSX mensual con UNA HOJA POR DÍA HÁBIL. Cada hoja lista TODOS los
Federal securities, incluidos los 3 Bund inflation-linked alemanes (Bund 14/21/15
"index.") con su yield REAL de mercado puro. De la hoja más reciente:

  * REAL10 = interpolación a 10 años de los linkers (real de mercado)
  * NOM10  = interpolación a 10 años de los Bund nominales (mismo snapshot)
  * BE10   = NOM10 - REAL10  (breakeven, internamente consistente)

Salida: data/RY_G8_EUR.csv  con cabecera idéntica al resto de tu pipeline:
        DATE,NOM10,REAL10,BE10   (DATE = YYYYMMDD, valores en %)

POR QUÉ
-------
Reemplaza el benchmark real MENSUAL del ECB (modelado) por un real DIARIO de
mercado. DE-core ~ EUR-real: el propio benchmark del ECB es una cesta de linkers
FR+DE; Alemania es el ancla AAA. Doble upgrade: mensual->diario y modelo->mercado.

CAVEATS HONESTOS (van marcados en la cabecera QUALITY del CSV)
--------------------------------------------------------------
  * DE-only, no la cesta agregada FR+DE del ECB.
  * 10Y va INTERPOLADO: no hay linker justo a 10y (el bracket es ~6.8y y ~19.8y).
  * Mercado de linkers alemán pequeño (~1% del volumen) -> algo de ruido.

AUTO-DESCUBRIMIENTO (clave: la URL cambia cada mes)
---------------------------------------------------
El enlace mensual lleva blob-id + hash + el YYYY-MM en el nombre, así que CAMBIA
cada mes. NUNCA se hardcodea: se raspa la página de listado y se elige el enlace
del mes vigente (fallback: el mes más reciente disponible).

FRESHNESS GATE
--------------
Si la hoja más reciente es más vieja que STALENESS_LIMIT_DAYS, sale con código !=0
(CI en rojo). Mata la clase de fallo "stale silencioso" igual que el de gilts.

USO
---
  python3 fetch_eur_real.py                 # produccion: descubre, descarga, escribe CSV
  python3 fetch_eur_real.py --test FILE.xlsx # valida el parseo contra un XLSX local (no escribe)
  python3 fetch_eur_real.py --dry-run        # descubre+descarga+parsea, imprime, NO escribe

Stdlib pura salvo openpyxl (ya lo usa tu scraper de gilts). Sin claves, sin auth.
"""

import sys
import os
import re
import ssl
import csv
import argparse
import datetime as dt
import urllib.request
import urllib.error

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("ERROR: falta openpyxl. En CI: pip install openpyxl\n")
    raise

# --------------------------------------------------------------------------- #
# CONFIG
# --------------------------------------------------------------------------- #
LISTING_URL = ("https://www.bundesbank.de/dynamic/action/en/service/"
               "federal-securities/prices-and-yields/810710/"
               "prices-and-yields-of-listed-federal-securities")
# Fallback de descubrimiento (mismo contenido, endpoint de busqueda ordenado por Latest)
LISTING_FALLBACK = "https://www.bundesbank.de/action/en/810710/bbksearch?sort=&query=*"

OUT_CSV = os.path.join("data", "RY_G8_EUR.csv")
TARGET_TENOR_Y = 10.0
STALENESS_LIMIT_DAYS = 7            # dias de calendario; hoja mas nueva no puede pasarse
QUALITY_TAG = "PROXY_DE_INTERP"     # DE-only + 10Y interpolado (honesto)
SOURCE_NOTE = "DE linkers interp to 10Y (Bundesbank; HICP basis)"

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

BASE = "https://www.bundesbank.de"
# Patron de los enlaces XLSX mensuales. Ancla en la RUTA (/resource/blob/...), por
# lo que casa tanto si el HTML trae el href relativo (/resource/...) como absoluto
# (https://www.bundesbank.de/resource/...). group(0)=ruta -> se le antepone BASE.
XLSX_RE = re.compile(
    r"/resource/blob/\d+/[0-9a-fA-F]+/[0-9A-F]+/"
    r"(\d{4})-(\d{2})-excel-data\.xlsx"
)
DATE_RE = re.compile(r"^\s*(\d{2})\.(\d{2})\.(\d{4})\s*$")   # DD.MM.YYYY


# --------------------------------------------------------------------------- #
# HTTP con fallback SSL relajado (red Bata con proxy de certificado self-signed)
# --------------------------------------------------------------------------- #
def _open(url, timeout=60):
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
    try:
        return urllib.request.urlopen(req, timeout=timeout)
    except urllib.error.URLError as e:
        reason = getattr(e, "reason", "")
        if "CERTIFICATE_VERIFY_FAILED" in str(reason) or isinstance(reason, ssl.SSLError):
            sys.stderr.write("[ssl] verify fallo -> reintento sin verificacion (red local)\n")
            ctx = ssl._create_unverified_context()
            return urllib.request.urlopen(req, timeout=timeout, context=ctx)
        raise


def http_text(url):
    with _open(url) as r:
        return r.read().decode("utf-8", "replace")


def http_bytes(url):
    with _open(url) as r:
        return r.read()


# --------------------------------------------------------------------------- #
# 1) DESCUBRIMIENTO del XLSX del mes vigente
# --------------------------------------------------------------------------- #
def discover_xlsx_url():
    """Raspa el listado y devuelve (url, (year, month)) del mes vigente; si no
    existe aun (rollover de mes), el mes mas reciente disponible."""
    html = ""
    for src in (LISTING_URL, LISTING_FALLBACK):
        try:
            html = http_text(src)
            if XLSX_RE.search(html):
                break
        except Exception as e:
            sys.stderr.write("[discover] aviso: fallo %s (%s)\n" % (src, e))
    matches = {}  # (y,m) -> url   (dedup; conserva el primero = top "Latest")
    for m in XLSX_RE.finditer(html):
        ym = (int(m.group(1)), int(m.group(2)))
        matches.setdefault(ym, BASE + m.group(0))
    if not matches:
        hint = ""
        if "excel-data" in html:
            hint = " (el HTML SI contiene 'excel-data' -> el regex no casa: revisa el patron)"
        elif "resource/blob" in html:
            hint = " (hay 'resource/blob' pero no '-excel-data.xlsx' -> ¿solo PDF en esta vista?)"
        else:
            hint = " (el HTML no contiene 'excel-data' ni 'resource/blob' -> contenido distinto/JS; len=%d)" % len(html)
        raise RuntimeError("DESCUBRIMIENTO FALLIDO: no encontre ningun enlace "
                           "*-excel-data.xlsx en el listado del Bundesbank." + hint)
    today = dt.date.today()
    cur = (today.year, today.month)
    if cur in matches:
        return matches[cur], cur
    latest = max(matches.keys())
    sys.stderr.write("[discover] aviso: mes vigente %04d-%02d no publicado aun; "
                     "uso el mas reciente %04d-%02d\n" % (cur[0], cur[1], latest[0], latest[1]))
    return matches[latest], latest


# --------------------------------------------------------------------------- #
# 2) PARSEO de la hoja mas reciente -> (sheet_date, NOM10, REAL10, BE10, detalle)
# --------------------------------------------------------------------------- #
def _parse_sheet_date(name):
    m = DATE_RE.match(name)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return dt.date(y, mo, d)
    except ValueError:
        return None


def _to_date(x):
    """Normaliza a datetime.date desde CUALQUIER forma que use el Bundesbank:
    objeto date/datetime (openpyxl data_only devuelve datetime en el layout nuevo
    2026-07), string ISO 'YYYY-MM-DD[ ...]', o string 'DD.MM.YYYY' (layout viejo).
    None si no parsea. (audit 2026-08: el rediseño de fin de junio paso las
    fechas de texto DD.MM.YYYY a objeto datetime -> este helper cubre ambos.)"""
    if x is None:
        return None
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x
    if isinstance(x, str):
        s = x.strip()
        m = DATE_RE.match(s)                              # DD.MM.YYYY (viejo)
        if m:
            try: return dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError: return None
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)      # ISO YYYY-MM-DD[...] (nuevo)
        if m:
            try: return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError: return None
    return None


def _is_linker(desc, isin):
    """Detecta un Bund ligado a inflacion — SOLO por descripcion.
    Layout nuevo (Bundesbank 2026-07): los linkers llevan prefijo 'DBRI' y sufijo
    'I/L' (p.ej. 'DBRI 0.1 04/15/33 I/L'); su yield en la tabla es el yield REAL.
    Los nominales normales, VERDES ('DBR … G') y GEMELOS ('DBR … TWI'/'… T') NO.
    IMPORTANTE (audit 2026-08, confirmado con --diagnose sobre el fichero de agosto):
    NO usar el prefijo de ISIN DE0001030 — esa familia incluye bonos VERDES
    NOMINALES (p.ej. DE0001030708 'DBR 0 08/15/30 G', yield 2.87%) que se colaban
    como linkers y contaminaban REAL10 con yields nominales. La descripcion es el
    unico discriminador fiable: 'DBRI'/'I/L' (viejo layout: 'index.')."""
    dl = (desc or '').lower()
    return any(k in dl for k in ('dbri', 'i/l', 'index', 'inflation', 'linker', 'ilb'))


# Layout posicional por defecto. Actualizado 2026-08 tras el rediseño Bundesbank:
# ISIN=0 · Coupon=1 · Description=2 · Maturity=3 · Yield=7 (antes 4/5/9). El
# auto-detector (autodetect_cols) recalibra igualmente si vuelve a cambiar.
POS_ISIN, POS_DESC, POS_MAT, POS_YLD = 0, 2, 3, 7


def _row_bond_at(row, sheet_date, desc_i, mat_i, yld_i):
    """Como _row_bond pero con indices de columna CONFIGURABLES. Devuelve
    (desc, residual_years, yield, is_linker) o None. Usado tanto por el layout
    posicional como por el auto-detector de columnas (fallback)."""
    n = len(row) if row is not None else 0
    if n <= max(desc_i, mat_i, yld_i):
        return None
    desc, mat, yld = row[desc_i], row[mat_i], row[yld_i]
    if not isinstance(desc, str):
        return None
    md = _to_date(mat)
    if md is None:
        return None
    try:
        y = float(yld)
    except (TypeError, ValueError):
        return None
    if not (-5.0 <= y <= 12.0):     # sanity band de yield
        return None
    resid = (md - sheet_date).days / 365.25
    if resid <= 0:
        return None
    isin = row[POS_ISIN] if n > POS_ISIN else ''
    return (desc.strip(), resid, y, _is_linker(desc, isin))


def _row_bond(row, sheet_date):
    """Layout posicional validado (desc=4, mat=5, yield=9). 'index.' -> linker."""
    if row is None or len(row) < 10:
        return None
    return _row_bond_at(row, sheet_date, POS_DESC, POS_MAT, POS_YLD)


def _extract_bonds(rows, sheet_date, desc_i, mat_i, yld_i):
    """Recorre las filas con un juego de indices dado y separa linkers/nominales."""
    linkers, nominals = [], []
    for row in rows:
        b = _row_bond_at(row, sheet_date, desc_i, mat_i, yld_i)
        if b is None:
            continue
        _desc, resid, y, is_linker = b
        (linkers if is_linker else nominals).append((resid, y, _desc))
    return linkers, nominals


def autodetect_cols(rows, sheet_date):
    """FALLBACK auto-calibrante (audit 2026-08).
    El parser posicional [4]/[5]/[9] se rompe en silencio si el Bundesbank
    inserta/mueve una columna (causa mas probable del feed congelado 2026-06-30,
    con descubrimiento + red ya verificados OK). La columna de VENCIMIENTO es
    inequivoca (muchas fechas DD.MM.YYYY futuras); la estructura interna de la
    tabla es estable, asi que el yield esta a un offset fijo del vencimiento
    (validado = +4) y la descripcion a -1. Se puntua cada combinacion candidata
    por nº de bonos validos y se elige la que da >=2 linkers y >=2 nominales.
    Devuelve (desc_i, mat_i, yld_i) o None. NUNCA se llama si el layout
    posicional ya funciona -> no puede regresionar el caso sano."""
    ncols = max((len(r) for r in rows if r), default=0)
    if ncols < 6:
        return None
    # 1) localizar la columna de vencimiento: la que tiene mas fechas FUTURAS
    #    (objeto datetime / ISO / DD.MM.YYYY — _to_date cubre las tres formas)
    matscore = [0] * ncols
    for r in rows:
        if not r:
            continue
        for j in range(min(len(r), ncols)):
            d = _to_date(r[j])
            if d is not None and d > sheet_date:
                matscore[j] += 1
    if not any(matscore):
        return None
    mat = matscore.index(max(matscore))
    # 2) confirmar offsets de yield/desc puntuando bonos validos
    best = None
    for dy in (4, 5, 3, 6, 2):          # offset yield vs vencimiento (validado +4)
        for dd in (-1, -2):             # offset descripcion
            yld, desc = mat + dy, mat + dd
            if yld >= ncols or desc < 0:
                continue
            lk, nm = _extract_bonds(rows, sheet_date, desc, mat, yld)
            if len(lk) >= 2 and len(nm) >= 2:
                score = len(lk) + len(nm)
                if best is None or score > best[0]:
                    best = (score, desc, mat, yld, len(lk), len(nm))
    if best:
        sys.stderr.write("[autodetect] layout recalibrado -> desc=%d mat=%d yld=%d "
                         "(%d linkers / %d nominales)\n"
                         % (best[1], best[2], best[3], best[4], best[5]))
        return best[1], best[2], best[3]
    sys.stderr.write("[autodetect] no encontre un layout con >=2 linkers y >=2 nominales "
                     "(mat col=%d). El Excel pudo cambiar de estructura mas alla de un "
                     "desplazamiento de columnas; usa --diagnose.\n" % mat)
    return None


def _interp(points, x):
    """Interpolacion lineal en (residual, yield). points: lista de (resid, yield).
    Si x cae fuera del rango, extrapola con los dos extremos (con aviso)."""
    pts = sorted(points)
    if len(pts) == 1:
        return pts[0][1]
    below = [p for p in pts if p[0] <= x]
    above = [p for p in pts if p[0] > x]
    if below and above:
        x0, y0 = below[-1]
        x1, y1 = above[0]
    elif not below:                 # x por debajo del minimo
        (x0, y0), (x1, y1) = pts[0], pts[1]
        sys.stderr.write("[interp] aviso: %.2fy por debajo del bracket; extrapolo\n" % x)
    else:                           # x por encima del maximo
        (x0, y0), (x1, y1) = pts[-2], pts[-1]
        sys.stderr.write("[interp] aviso: %.2fy por encima del bracket; extrapolo\n" % x)
    if x1 == x0:
        return y0
    return y0 + (x - x0) * (y1 - y0) / (x1 - x0)


BRACKET_TOL_Y = 1.5   # backfill: 10Y puede extrapolarse como mucho 1.5y mas alla del
                      # span de linkers; mas alla, se descarta la fila (real no fiable).


def compute_sheet(ws, sheet_date, strict=True, require_bracket=False,
                  bracket_tol=BRACKET_TOL_Y):
    """Extrae linkers + nominales de UNA hoja e interpola a 10Y.
    strict=True  -> lanza si faltan datos (uso diario: la ultima hoja siempre los tiene).
    strict=False -> devuelve None si faltan datos (uso backfill: salta ese dia).
    require_bracket=True -> ademas exige que 10Y caiga dentro del span de linkers
                            (+/- bracket_tol); si no, None (real seria extrapolacion
                            grande, no fiable). Guardarrail del backfill."""
    rows = list(ws.iter_rows(values_only=True))
    # 1) intento con el layout posicional validado (comportamiento sin cambios)
    linkers, nominals = _extract_bonds(rows, sheet_date, POS_DESC, POS_MAT, POS_YLD)

    # 2) FALLBACK auto-calibrante (audit 2026-08): si el posicional ya no da datos
    #    suficientes -> el Excel cambio de columnas. Se recalibra sobre esta misma
    #    hoja y se reintenta. Solo entra en juego cuando el posicional falla.
    if len(linkers) < 2 or len(nominals) < 2:
        cols = autodetect_cols(rows, sheet_date)
        if cols is not None:
            linkers, nominals = _extract_bonds(rows, sheet_date, cols[0], cols[1], cols[2])

    if len(linkers) < 2 or len(nominals) < 2:
        if strict:
            raise RuntimeError("PARSEO FALLIDO: hoja %s con %d linkers / %d nominales "
                               "(esperaba >=2 de cada; ni el layout posicional ni el "
                               "auto-detector encontraron suficientes bonos)."
                               % (sheet_date.isoformat(), len(linkers), len(nominals)))
        return None

    if require_bracket:
        lr = sorted(r for (r, _, _) in linkers)
        if not (lr[0] - bracket_tol <= TARGET_TENOR_Y <= lr[-1] + bracket_tol):
            return None   # 10Y demasiado lejos del span de linkers -> real no fiable

    real10 = _interp([(r, y) for (r, y, _) in linkers], TARGET_TENOR_Y)
    nom10 = _interp([(r, y) for (r, y, _) in nominals], TARGET_TENOR_Y)
    return {
        "date": sheet_date,
        "sheet": sheet_date.strftime("%d.%m.%Y"),
        "NOM10": round(nom10, 4),
        "REAL10": round(real10, 4),
        "BE10": round(nom10 - real10, 4),
        "linkers": sorted(linkers),
        "n_nominals": len(nominals),
    }


def _sheets_by_date(xlsx_bytes_or_path):
    wb = load_workbook(xlsx_bytes_or_path, read_only=True, data_only=True)
    dated = [(_parse_sheet_date(s), s) for s in wb.sheetnames]
    dated = [(d, s) for (d, s) in dated if d is not None]
    if not dated:
        raise RuntimeError("PARSEO FALLIDO: ninguna hoja con nombre DD.MM.YYYY "
                           "(layout inesperado).")
    return wb, sorted(dated, key=lambda t: t[0])


def parse_workbook(xlsx_bytes_or_path):
    """Diario: elige la hoja con fecha MAXIMA e interpola a 10Y (strict)."""
    wb, dated = _sheets_by_date(xlsx_bytes_or_path)
    sheet_date, sheet_name = dated[-1]
    return compute_sheet(wb[sheet_name], sheet_date, strict=True)


def parse_all_sheets(xlsx_bytes_or_path, require_bracket=True):
    """Backfill: una fila por hoja (dia habil), saltando las que no dan un REAL10
    fiable (guardarrail de bracket). Devuelve lista de results ordenada por fecha."""
    wb, dated = _sheets_by_date(xlsx_bytes_or_path)
    out = []
    for sheet_date, sheet_name in dated:
        r = compute_sheet(wb[sheet_name], sheet_date, strict=False,
                          require_bracket=require_bracket)
        if r is not None:
            out.append(r)
    return out


# --------------------------------------------------------------------------- #
# 3) ESCRITURA idempotente de RY_G8_EUR.csv  (upsert por DATE)
# --------------------------------------------------------------------------- #
HEADER_COMMENT = "# QUALITY=%s | %s | generated by fetch_eur_real.py" % (QUALITY_TAG, SOURCE_NOTE)
COLS = ["DATE", "NOM10", "REAL10", "BE10"]


def _result_row(result):
    d = result["date"].strftime("%Y%m%d")
    return d, [d, "%g" % result["NOM10"], "%g" % result["REAL10"], "%g" % result["BE10"]]


def _upsert(new_rows, path=OUT_CSV):
    """new_rows: dict {YYYYMMDD: [d,nom,real,be]}. Lee existente, fusiona, reescribe.
    csv.reader maneja comillas/terminadores; last-write-wins por fecha."""
    rows = {}
    if os.path.exists(path):
        with open(path, "r", newline="") as f:
            for parts in csv.reader(f):
                if not parts:
                    continue
                head = parts[0].strip()
                if head.startswith("#") or head == "DATE":
                    continue
                if len(parts) >= 4 and head.isdigit():
                    rows[head] = [p.strip() for p in parts[:4]]
    rows.update(new_rows)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", newline="") as f:
        f.write(HEADER_COMMENT + "\n")
        w = csv.writer(f, lineterminator="\n")
        w.writerow(COLS)
        for k in sorted(rows.keys()):
            w.writerow(rows[k])
    return len(rows)


def write_csv(result, path=OUT_CSV):
    d, row = _result_row(result)
    _upsert({d: row}, path)
    return d


def write_rows(results, path=OUT_CSV):
    """Backfill: upsert idempotente de muchos results de una vez.
    Devuelve (filas_nuevas_aportadas, total_filas_en_csv)."""
    nr = dict(_result_row(r) for r in results)
    total = _upsert(nr, path)
    return len(nr), total


# --------------------------------------------------------------------------- #
# 4) FRESHNESS GATE
# --------------------------------------------------------------------------- #
def freshness_gate(result):
    age = (dt.date.today() - result["date"]).days
    if age > STALENESS_LIMIT_DAYS:
        sys.stderr.write(
            "FRESHNESS GATE FALLO: la hoja mas reciente es %s (%d dias), "
            "limite %d. CI en rojo a proposito.\n"
            % (result["date"].isoformat(), age, STALENESS_LIMIT_DAYS))
        return False
    return True


def _print_summary(result):
    print("Hoja mas reciente : %s  (date=%s)" % (result["sheet"], result["date"].isoformat()))
    print("Linkers (resid y, real%):")
    for r, y, desc in result["linkers"]:
        print("   %-22s %5.2fy  real %.3f%%" % (desc, r, y))
    print("Nominales usados  : %d puntos de curva" % result["n_nominals"])
    print("-> NOM10  = %.3f%%" % result["NOM10"])
    print("-> REAL10 = %.3f%%  (interpolado a 10Y)" % result["REAL10"])
    print("-> BE10   = %.3f%%  (= NOM10 - REAL10)" % result["BE10"])


# --------------------------------------------------------------------------- #
# MAIN
# --------------------------------------------------------------------------- #
def diagnose(xlsx_bytes_or_path):
    """Vuelca la estructura de la hoja mas reciente para depurar un cambio de
    layout: nº de hojas, cabeceras/columnas por indice, que ve el parser
    posicional y que ve el auto-detector. Uso: --diagnose (red) o --diagnose FILE."""
    wb, dated = _sheets_by_date(xlsx_bytes_or_path)
    sheet_date, sheet_name = dated[-1]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    ncols = max((len(r) for r in rows if r), default=0)
    print("Hojas con fecha : %d  ·  ultima = %s (%s)" % (len(dated), sheet_name, sheet_date.isoformat()))
    print("Columnas (max)  : %d" % ncols)
    print("\nTODAS las filas con datos (indice:valor, truncado) · ⟵LINKER? marca los que")
    print("el detector cree ligados a inflacion (por descripcion o ISIN DE0001030…):")
    for r in rows:
        if not r or all(c is None or c == "" for c in r):
            continue
        cells = " | ".join("%d:%s" % (j, str(r[j])[:18]) for j in range(len(r)) if r[j] not in (None, ""))
        isin = r[POS_ISIN] if len(r) > POS_ISIN else ''
        desc = r[POS_DESC] if len(r) > POS_DESC else ''
        mark = ' ⟵LINKER?' if (isinstance(desc, str) and _is_linker(desc, isin if isinstance(isin, str) else '')) else ''
        print("  " + cells[:210] + mark)
    lk, nm = _extract_bonds(rows, sheet_date, POS_DESC, POS_MAT, POS_YLD)
    print("\nPosicional [desc=%d mat=%d yld=%d] -> %d linkers / %d nominales" % (POS_DESC, POS_MAT, POS_YLD, len(lk), len(nm)))
    if lk:
        print("  linkers detectados:")
        for r_, y_, d_ in sorted(lk):
            print("    %-24s %5.2fy  real %.3f%%" % (d_, r_, y_))
    cols = autodetect_cols(rows, sheet_date)
    if cols:
        lk2, nm2 = _extract_bonds(rows, sheet_date, cols[0], cols[1], cols[2])
        print("Auto-detector    -> desc=%d mat=%d yld=%d  ·  %d linkers / %d nominales"
              % (cols[0], cols[1], cols[2], len(lk2), len(nm2)))
    else:
        print("Auto-detector    -> sin layout valido (revisa el volcado de columnas arriba)")
    return 0


def main():
    ap = argparse.ArgumentParser(description="EUR(DE) real 10Y diario desde Bundesbank linkers")
    ap.add_argument("--test", metavar="XLSX", help="valida el parseo contra un XLSX local (no escribe, no freshness)")
    ap.add_argument("--dry-run", action="store_true", help="descubre+descarga+parsea, imprime, NO escribe")
    ap.add_argument("--diagnose", nargs="?", const="__net__", metavar="XLSX",
                    help="vuelca la estructura de la hoja mas reciente (red, o un XLSX local) para depurar layout")
    args = ap.parse_args()

    # Modo diagnose: estructura de columnas (local o descargando el mes vigente)
    if args.diagnose:
        if args.diagnose == "__net__":
            url, ym = discover_xlsx_url()
            sys.stderr.write("[discover] XLSX %04d-%02d -> %s\n" % (ym[0], ym[1], url))
            import io
            return diagnose(io.BytesIO(http_bytes(url)))
        return diagnose(args.diagnose)

    # Modo test: parsea un fichero local, sin red, sin escribir, sin gate
    if args.test:
        result = parse_workbook(args.test)
        _print_summary(result)
        print("\n[--test] OK (no se escribio CSV, no se aplico freshness gate).")
        return 0

    # Produccion / dry-run
    url, ym = discover_xlsx_url()
    sys.stderr.write("[discover] XLSX %04d-%02d -> %s\n" % (ym[0], ym[1], url))
    data = http_bytes(url)
    import io
    result = parse_workbook(io.BytesIO(data))
    _print_summary(result)

    if not freshness_gate(result):
        return 1

    if args.dry_run:
        print("\n[--dry-run] OK (no se escribio CSV).")
        return 0

    datestr = write_csv(result)
    print("\nEscrito %s -> fila %s" % (OUT_CSV, datestr))
    return 0


if __name__ == "__main__":
    sys.exit(main())
